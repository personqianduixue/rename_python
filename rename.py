import os
from openpyxl import load_workbook
import re
import warnings
import argparse
parser = argparse.ArgumentParser()
parser.add_argument('-student_data',default='./Students.xlsx')
parser.add_argument('-start_str',type=str,default='集成1901')
parser.add_argument('-end_str',type=str,default='第一次作业')
args = parser.parse_args()
warnings.filterwarnings('ignore')

# 根据字典的值value获得该值对应的key
def get_dict_key(dic, value):
    key = list(dic.keys())[list(dic.values()).index(value)]
    return key

look_up_table_path=args.student_data
look_up_table_row_start=2
# look_up_table_row_number=27
name_number_dict={}
look_up_table_excel=load_workbook(look_up_table_path)
look_up_table_all_sheet=look_up_table_excel.get_sheet_names()
look_up_table_sheet=look_up_table_excel.get_sheet_by_name(look_up_table_all_sheet[0])
look_up_table_row_number=look_up_table_sheet.max_row
print(look_up_table_row_number)
for i in range(look_up_table_row_start,look_up_table_row_start+look_up_table_row_number):
    number=str(look_up_table_sheet.cell(i,2).value)
    name=look_up_table_sheet.cell(i,1).value
    if number is not 'None' and name is not None:
        name_number_dict[name]=number


name_number_list=list(name_number_dict.values())+(list(name_number_dict.keys()))
name_list = list(name_number_dict.keys())
start_str = args.start_str
end_str = args.end_str
paths = os.listdir('.')
for path in paths:
    new_path = path
    for name_number in name_number_list:
        name_number = str(name_number)
        if name_number in path and name_number:
            if bool(re.search(r'\d', name_number)):
                number = name_number
                name = get_dict_key(name_number_dict, name_number)
            else:
                name = name_number
                number = name_number_dict[name_number]
            print(name)
            if name in name_list:
                name_list.remove(name)
            if os.path.isdir(path):
                new_path = start_str + '-' + number + '-' + name + '-' + end_str
            else:
                ext = os.path.splitext(path)[-1]
                new_path = start_str + '-' + number + '-' + name + '-' +  end_str +ext
            break
    print(new_path)
    if path != new_path:
        os.rename(path, new_path)
file = './未交作业名单.txt'
if os.path.exists(file):
    os.remove(file)
print('未交作业名单：',name_list)
with open(file,'w') as f:
    f.write(str(name_list))



